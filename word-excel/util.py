from openpyxl import load_workbook
from openpyxl.utils import get_column_letter 
import pandas as pd

# 删除空行
def remove(sheet, row):
    for cell in row:
        if cell.value != None:
                return
    sheet.delete_rows(row[0].row, 1)

# 调整列宽
def reset_col(filename):
    wb=load_workbook(filename)
    for sheet in wb.sheetnames:
        ws=wb[sheet]
        df=pd.read_excel(filename,sheet).fillna('-')
        if not df.empty:
            df.loc[len(df)]=list(df.columns)						#把标题行附加到最后一行
            for col in df.columns:				
                index=list(df.columns).index(col)					#列序号
                letter=get_column_letter(index+1)					#列字母
                collen=df[col].apply(lambda x:len(str(x).encode())).max()	#获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
                ws.column_dimensions[letter].width=collen*1+2		#也就是列宽为最大长度*1.2 可以自己调整
 
    wb.save(filename)